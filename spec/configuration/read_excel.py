import openpyxl
import json

start_row = 2
max_row = 5
start_col = 3
max_col = 10


wb = openpyxl.load_workbook("采集点位列表.xlsx")
sh = wb['Sheet1']


#table_name_list = []
#for c in range (start_col, start_col+max_col):
#    ce = sh.cell(row = 1,column = c)   
#    print(ce.value, end=' ')
#    table_name_list.append(ce.value)
#

fjson =  open('dataitem.json','w')
json_header_string = '''{\n"data-item-list": [\n'''

json_format_str = '''{\n"address": "%s",\n"data-transform": "%s",\n"data-type": "%s",\n"modbus-id": "%s",\n"modbus-station-id": "%s",\n"operation": "%s",\n "value": %s,\n"widget": {\n"widget-id": "%s",\n"widget-item": "%s",\n"widget-type": "%s"\n}\n},\n'''

json_string = ""
json_string = json_string + json_header_string
for r in range (start_row, start_row + max_row):
    value_list = []
    tmp_str = ""
    for c in range (start_col, start_col + max_col):
        ce = sh.cell(row = r,column = c)   
        #print(ce.value, end=' ')
        value_list.append(ce.value)

    #print(value_list)
    tmp_str = json_format_str %tuple(value_list)
    #print(tmp_str)
    json_string += tmp_str;
wb.close()

json_end_string = ''']\n}\n'''
json_string = json_string[:len(json_string)-2] + json_end_string
fjson.write(json.dumps(json.loads(json_string), sort_keys=True, indent=2)) 
fjson.close()
